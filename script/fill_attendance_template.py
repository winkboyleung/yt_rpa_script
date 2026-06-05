"""
按日增量填写考勤模板（默认：昨天）。
一次运行传入：1 个模板 + 1 份打卡记录；按文件名识别中介 / 办公室(含正式员工) 版式。

第 1 步请先跑 rpa_4_clock_in.py 生成整月异常表（方案 B：全月扫描）。
"""
import argparse
import os
import re
import sys
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.holiday_checker import is_holiday_or_weekend
from utils import workday_overtime
from utils.workday_overtime import refresh_workday_six_punch_employees_from_df
from utils.agency_attendance import get_agency_employee_keys
from utils.night_shift import collect_missing_night_start_dates
from utils.attendance_cells import (
    build_employee_anomaly_maps,
    merge_missing_night_start_anomalies,
    compute_attendance_cell,
    compute_overtime_cell,
    compute_agency_template_daily_cell,
)

FILES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "files"
)

# ── PyCharm 右击「运行」默认配置（改这里即可，不必配命令行参数）──
DEFAULT_TEMPLATE = os.path.join(FILES_DIR, "6月中介考勤模版.xlsx")
# 换模板时改上一行，或取消下面某一行的注释：
# DEFAULT_TEMPLATE = os.path.join(FILES_DIR, "6月正式员工考勤模版.xlsx")
# DEFAULT_TEMPLATE = os.path.join(FILES_DIR, "6月中介考勤模版.xlsx")

DEFAULT_PUNCH = os.path.join(FILES_DIR, "6月打卡.xls")
DEFAULT_ANOMALY = os.path.join(FILES_DIR, "6月打卡异常.xlsx")

DEFAULT_LOOKBACK_DAYS = 4  # 1=仅昨天，2=前天+昨天，3=大前天+前天+昨天
# ── 以上默认值 ──

TEMPLATE_KIND_AGENCY = "agency"
TEMPLATE_KIND_OFFICE = "office"

# 办公室 / 正式员工模版
OFFICE_DAY_HEADER_ROW = 3
OFFICE_DATA_START_ROW = 5
OFFICE_NAME_COL = 2

# 中介模版
AGENCY_DAY_HEADER_ROW = 2
AGENCY_DATA_START_ROW = 4
AGENCY_NAME_COL = 3
AGENCY_SKIP_NAMES = frozenset({"姓名", "宵夜"})


def detect_template_kind(template_path):
    """根据文件名判断模版版式：中介 / 办公室(含正式员工)。"""
    name = os.path.basename(template_path)
    if "中介" in name:
        return TEMPLATE_KIND_AGENCY
    return TEMPLATE_KIND_OFFICE


def read_punch_df(path):
    try:
        return pd.read_excel(path, engine="xlrd")
    except Exception:
        return pd.read_excel(path, engine="openpyxl")


def _parse_day_columns(ws, header_row, col_start=4, col_end=50):
    day_to_col = {}
    for col in range(col_start, col_end):
        val = ws.cell(header_row, col).value
        if isinstance(val, int) and 1 <= val <= 31:
            day_to_col[val] = col
    return day_to_col


def parse_office_template_layout(ws):
    """{日号: 列号}, {姓名: (出勤行, 加班行)}"""
    day_to_col = _parse_day_columns(ws, OFFICE_DAY_HEADER_ROW)
    name_to_rows = {}
    row = OFFICE_DATA_START_ROW
    max_row = ws.max_row or 500
    while row <= max_row:
        name = ws.cell(row, OFFICE_NAME_COL).value
        label = ws.cell(row, 3).value
        if name and str(label).strip() == "正常出勤":
            name = str(name).strip()
            name_to_rows[name] = (row, row + 1)
            row += 2
            continue
        if name is None and label is None and row > OFFICE_DATA_START_ROW + 4:
            break
        row += 1
    return day_to_col, name_to_rows


def parse_agency_template_layout(ws):
    """{日号: 列号}, {姓名: 行号}"""
    day_to_col = _parse_day_columns(ws, AGENCY_DAY_HEADER_ROW)
    name_to_row = {}
    max_row = ws.max_row or 800
    for row in range(AGENCY_DATA_START_ROW, max_row + 1):
        name = ws.cell(row, AGENCY_NAME_COL).value
        if not name:
            continue
        name = str(name).strip()
        if name in AGENCY_SKIP_NAMES:
            continue
        name_to_row[name] = row
    return day_to_col, name_to_row


def infer_year_month(ws, template_path, punch_df):
    for text in (ws.cell(1, 1).value, os.path.basename(template_path)):
        if not text:
            continue
        s = str(text)
        m = re.search(r"(\d{4})年(\d{1,2})月", s)
        if m:
            return int(m.group(1)), int(m.group(2))
    start = punch_df["日期"].min()
    return start.year, start.month


def target_dates(year, month, reference_date, lookback_days):
    month_start = datetime(year, month, 1).date()
    if month == 12:
        month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)

    dates = []
    for i in range(1, lookback_days + 1):
        d = reference_date - timedelta(days=i)
        if month_start <= d <= month_end:
            dates.append(d)
    return sorted(dates)


def apply_cell(ws, row, col, value, comment_text, number_format=None):
    cell = ws.cell(row, col)
    cell.value = value
    if number_format is not None:
        cell.number_format = number_format
    if comment_text:
        cell.comment = Comment(comment_text, "系统")
    elif cell.comment:
        cell.comment = None


# 中介模版部分日期列带自定义格式 "夜"##，写入 0 时 Excel 会显示成「夜」
AGENCY_DAY_CELL_NUMBER_FORMAT = "General"


def _load_employee_context(name, punch_df, anomaly_df, month_start, month_end, agency_keys):
    emp_records = punch_df[punch_df["姓名"] == name]
    if emp_records.empty:
        emp_id = ""
        emp_punches_by_date = {}
        emp_punch_dates = set()
    else:
        emp_id = emp_records.iloc[0]["编号"]
        emp_punch_dates = set(emp_records["日期"])
        emp_punches_by_date = {}
        for d, group in emp_records.groupby("日期"):
            emp_punches_by_date[d] = sorted(group["打卡时间"].tolist())

    emp_anomalies, emp_anomaly_details = build_employee_anomaly_maps(
        name, anomaly_df
    )
    missing_night_start = {}
    is_four_punch = name in workday_overtime.TARGET_WORKDAY_SIX_PUNCH_EMPLOYEES
    if is_four_punch:
        missing_night_start = collect_missing_night_start_dates(
            emp_punches_by_date, month_start, month_end
        )
        merge_missing_night_start_anomalies(
            name, emp_anomalies, emp_anomaly_details, missing_night_start
        )

    is_agency = (name, emp_id) in agency_keys if emp_id else False
    return {
        "emp_id": emp_id,
        "emp_punch_dates": emp_punch_dates,
        "emp_punches_by_date": emp_punches_by_date,
        "emp_anomalies": emp_anomalies,
        "emp_anomaly_details": emp_anomaly_details,
        "missing_night_start": missing_night_start,
        "is_four_punch": is_four_punch,
        "is_agency": is_agency,
    }


def _fill_office_template(
    ws, name_to_rows, day_to_col, dates_to_fill, punch_df, anomaly_df,
    month_start, month_end, agency_keys,
):
    filled_cells = 0
    for name, (att_row, ot_row) in name_to_rows.items():
        ctx = _load_employee_context(
            name, punch_df, anomaly_df, month_start, month_end, agency_keys
        )
        for punch_date in dates_to_fill:
            col = day_to_col.get(punch_date.day)
            if not col:
                continue

            att_val, att_comment = compute_attendance_cell(
                punch_date,
                ctx["emp_punch_dates"],
                ctx["emp_punches_by_date"],
                ctx["emp_anomalies"],
                ctx["emp_anomaly_details"],
                ctx["is_agency"],
                ctx["is_four_punch"],
                name=name,
            )
            if att_val is not None:
                apply_cell(ws, att_row, col, att_val, att_comment)
                filled_cells += 1

            ot_val, ot_comment = compute_overtime_cell(
                name,
                ctx["emp_id"],
                punch_date,
                ctx["emp_punches_by_date"],
                ctx["missing_night_start"],
                ctx["is_agency"],
                ctx["is_four_punch"],
            )
            if ot_val is not None:
                apply_cell(ws, ot_row, col, ot_val, ot_comment)
                filled_cells += 1
            elif not is_holiday_or_weekend(punch_date):
                apply_cell(ws, ot_row, col, None, None)

    return filled_cells


def _fill_agency_template(
    ws, name_to_row, day_to_col, dates_to_fill, punch_df, anomaly_df,
    month_start, month_end, agency_keys,
):
    filled_cells = 0
    for name, row in name_to_row.items():
        ctx = _load_employee_context(
            name, punch_df, anomaly_df, month_start, month_end, agency_keys
        )
        for punch_date in dates_to_fill:
            col = day_to_col.get(punch_date.day)
            if not col:
                continue

            val, comment = compute_agency_template_daily_cell(
                punch_date,
                ctx["emp_punches_by_date"],
                ctx["emp_anomalies"],
                ctx["emp_anomaly_details"],
            )
            apply_cell(
                ws, row, col, val, comment,
                number_format=AGENCY_DAY_CELL_NUMBER_FORMAT,
            )
            filled_cells += 1

    return filled_cells


def fill_attendance_template(
    template_path,
    punch_path,
    anomaly_path,
    reference_date=None,
    lookback_days=1,
    output_path=None,
):
    reference_date = reference_date or datetime.now().date()
    output_path = output_path or template_path
    kind = detect_template_kind(template_path)

    punch_df = read_punch_df(punch_path)
    punch_df["日期时间"] = pd.to_datetime(punch_df["日期时间"])
    punch_df["日期"] = punch_df["日期时间"].dt.date
    punch_df["打卡时间"] = punch_df["日期时间"].dt.time

    refresh_workday_six_punch_employees_from_df(punch_df)
    agency_keys = get_agency_employee_keys(punch_df)

    anomaly_df = pd.DataFrame()
    if os.path.exists(anomaly_path):
        anomaly_df = pd.read_excel(anomaly_path)
        anomaly_df["日期"] = pd.to_datetime(anomaly_df["日期"]).dt.date

    wb = load_workbook(template_path)
    ws = wb.active

    if kind == TEMPLATE_KIND_AGENCY:
        day_to_col, name_map = parse_agency_template_layout(ws)
        layout_label = "中介（单行·出考勤工时）"
    else:
        day_to_col, name_map = parse_office_template_layout(ws)
        layout_label = "办公室/正式员工（双行）"

    year, month = infer_year_month(ws, template_path, punch_df)
    dates_to_fill = target_dates(year, month, reference_date, lookback_days)

    if not dates_to_fill:
        print(f"无落在 {year}年{month}月 内的目标日期（参考日 {reference_date}，回溯 {lookback_days} 天）")
        return

    month_start = datetime(year, month, 1).date()
    days_in_month = pd.Period(f"{year}-{month}").days_in_month
    month_end = datetime(year, month, days_in_month).date()

    print(f"模板: {template_path}")
    print(f"版式: {layout_label}")
    print(f"模板员工: {len(name_map)} 人")
    print(f"参考日: {reference_date}，填写日期: {[d.isoformat() for d in dates_to_fill]}")

    if kind == TEMPLATE_KIND_AGENCY:
        filled_cells = _fill_agency_template(
            ws, name_map, day_to_col, dates_to_fill, punch_df, anomaly_df,
            month_start, month_end, agency_keys,
        )
    else:
        filled_cells = _fill_office_template(
            ws, name_map, day_to_col, dates_to_fill, punch_df, anomaly_df,
            month_start, month_end, agency_keys,
        )

    wb.save(output_path)
    print(f"已写入 {filled_cells} 个单元格")
    print(f"已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="按日填写考勤模板（办公室/正式/中介，由文件名自动识别）"
    )
    parser.add_argument("--template", default=DEFAULT_TEMPLATE, help="考勤模板路径")
    parser.add_argument("--punch", default=DEFAULT_PUNCH, help="打卡记录路径")
    parser.add_argument("--anomaly", default=DEFAULT_ANOMALY, help="异常表路径")
    parser.add_argument(
        "--days",
        type=int,
        default=DEFAULT_LOOKBACK_DAYS,
        help="回溯天数，默认见 DEFAULT_LOOKBACK_DAYS",
    )
    parser.add_argument(
        "--reference-date",
        default=None,
        help="参考日期 YYYY-MM-DD，默认今天",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="输出路径，默认覆盖原模板",
    )
    args = parser.parse_args()
    ref = (
        datetime.strptime(args.reference_date, "%Y-%m-%d").date()
        if args.reference_date
        else datetime.now().date()
    )
    fill_attendance_template(
        args.template,
        args.punch,
        args.anomaly,
        reference_date=ref,
        lookback_days=args.days,
        output_path=args.output,
    )


if __name__ == "__main__":
    main()

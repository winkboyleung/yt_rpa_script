import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.comments import Comment
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils.holiday_checker import is_holiday_or_weekend, is_workday
from utils import workday_overtime
from utils.workday_overtime import (
    TARGET_WORKDAY_OVERTIME_EMPLOYEES,
    calc_workday_overtime,
    calc_four_punch_department_rest_overtime,
    refresh_workday_six_punch_employees_from_df,
)
from utils.agency_attendance import (
    get_agency_employee_keys,
    calc_agency_hours_for_shift_end_date,
)
from utils.night_shift import collect_missing_night_start_dates

# 文件路径
PUNCH_FILE = "/Applications/ramsey_leung_files/all_files_from_redmi/yt_rpa_script/files/5月办公室打卡.xls"
ANOMALY_FILE = "/Applications/ramsey_leung_files/all_files_from_redmi/yt_rpa_script/files/五月打卡异常.xlsx"
OUTPUT_DIR = "/Applications/ramsey_leung_files/all_files_from_redmi/yt_rpa_script/files"


def _time_to_minutes(t):
    return t.hour * 60 + t.minute


def _round_clock_in_forward(t):
    """
    上班卡取整到半点：
    - 早上早于8:30的打卡，一律按8:30计（7:45、7:59→8:30）
    - 若距离“上一半点”小于5分钟，则仍归到上一半点（9:34→9:30，10:03→10:00）
    - 否则归到下一半点（8:25→8:30，9:55→10:00）
    """
    mins = _time_to_minutes(t)
    if mins < 8 * 60 + 30:
        return 8 * 60 + 30
    remainder = mins % 30
    if remainder == 0:
        return mins
    if remainder < 5:
        return mins - remainder
    return (mins // 30 + 1) * 30


def _round_clock_out_backward(t):
    """下班卡向后取整到半点（17:25→17:00，18:11→18:00）"""
    mins = _time_to_minutes(t)
    return (mins // 30) * 30


def calc_weekend_overtime_two_punches(clock_in, clock_out):
    """
    周末/节假日、两次打卡的加班工时。
    下班(向后取整) - 上班(向前取整)；
    >5小时减1小时午餐（上班>=11:00不扣）；
    或者原始上班<11:00且原始下班>14:00时，也减1小时午餐（上班>=11:00不扣）。
    """
    start = _round_clock_in_forward(clock_in)
    end = _round_clock_out_backward(clock_out)
    if end <= start:
        return None
    hours = (end - start) / 60.0
    # 上班 >= 11:00 视为已吃过午饭，不再扣午餐1小时
    need_lunch_deduction = _time_to_minutes(clock_in) < 11 * 60 and (
        hours > 5
        or _time_to_minutes(clock_out) > 14 * 60
    )
    if need_lunch_deduction:
        hours -= 1
    # 下班时间 > 19:30，额外减0.5小时晚餐
    if _time_to_minutes(clock_out) > 19 * 60 + 30:
        hours -= 0.5
    return hours


def _format_overtime_hours(hours):
    # 统一格式化，避免出现 "8." 这类尾点显示
    text = f"{round(hours, 2):.2f}".rstrip("0")
    if text.endswith("."):
        text = text[:-1]
    return text


def _stats_hours_value(total):
    """统计区工时：无加班为 0，有则与日历格相同格式。"""
    if not total:
        return 0
    return _format_overtime_hours(total)


STATS_COL_START = 35  # AI列
STATS_COL_END = 57    # BE列

# 正常出勤统计列
COL_STATS_ACTUAL = 35       # AI 实际出勤
COL_STATS_COMP_LEAVE = 36   # AJ 调休
COL_STATS_ANNUAL_LEAVE = 37   # AK 年休假
COL_STATS_ATTEND_SUM = 38     # AL 合计

# 缺勤(天)统计列
COL_STATS_PERSONAL_LEAVE = 39   # AM 事假
COL_STATS_SICK_LEAVE = 40       # AN 病假
COL_STATS_FAMILY_LEAVE = 41     # AO 婚丧产
COL_STATS_WORK_INJURY = 42      # AP 工伤
COL_STATS_ABSENT = 43           # AQ 旷工

# 加班工时(H)统计列
COL_STATS_OT_WEEKDAY = 44       # AR 平时
COL_STATS_OT_REST = 45          # AS 休息日(周末+节假日)
COL_STATS_OT_HOLIDAY = 46       # AT 节假日

# 统计区列宽：竖排表头列宜窄且同组等宽，避免 AN/AU/BA 等过宽导致参差不齐
STATS_COL_WIDTHS = {
    "AI": 6.0, "AJ": 5.0, "AK": 4.75, "AL": 6.25,
    "AM": 4.0, "AN": 4.0, "AO": 4.0, "AP": 4.0, "AQ": 4.0,
    "AR": 4.25, "AS": 4.25, "AT": 4.25,
    "AU": 4.25, "AV": 5.5,
    "AW": 4.75, "AX": 5.75,
    "AY": 5.875, "AZ": 4.5, "BA": 4.75,
    "BB": 10.25, "BC": 6.25, "BD": 10.75, "BE": 4.75,
}


def _style_cell(cell, font, alignment, border):
    cell.font = font
    cell.alignment = alignment
    cell.border = border


def _thin_border_sides(top=False, bottom=False, left=False, right=False):
    side = Side(style="thin")
    none = Side()
    return Border(
        top=side if top else none,
        bottom=side if bottom else none,
        left=side if left else none,
        right=side if right else none,
    )


def _full_thin_border():
    return _thin_border_sides(top=True, bottom=True, left=True, right=True)


def _apply_horizontal_merge_borders(ws, row, min_col, max_col):
    """单行横向合并：顶底边铺满，左右边仅在最外侧单元格。"""
    for col in range(min_col, max_col + 1):
        ws.cell(row, col).border = _thin_border_sides(
            top=True,
            bottom=True,
            left=(col == min_col),
            right=(col == max_col),
        )


def _apply_vertical_merge_borders(ws, col, min_row, max_row):
    """单列纵向合并：上格四边，下格底+左右。"""
    for row in range(min_row, max_row + 1):
        if row == min_row:
            ws.cell(row, col).border = _full_thin_border()
        else:
            ws.cell(row, col).border = _thin_border_sides(bottom=True, left=True, right=True)


def _apply_merge_range_borders(ws, merge_range):
    min_col, min_row, max_col, max_row = range_boundaries(merge_range)
    if min_row == max_row:
        _apply_horizontal_merge_borders(ws, min_row, min_col, max_col)
    elif min_col == max_col:
        _apply_vertical_merge_borders(ws, min_col, min_row, max_row)
    else:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row, col).border = _full_thin_border()


def _vertical_header_align():
    """Excel 竖排文字（自上而下）。"""
    return Alignment(horizontal="center", vertical="center", text_rotation=255, wrap_text=True)


def _setup_stats_headers(ws, workdays, header_font, center_align):
    """右侧统计区表头（仅样式，不填数据）。"""
    ws.row_dimensions[3].height = 47.25
    ws.row_dimensions[4].height = 69.75
    vertical_align = _vertical_header_align()
    unit_bottom_align = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    full_border = _full_thin_border()

    stats_merges_row3 = [
        ("AI3:AL3", f"正常出勤({workdays}天)"),
        ("AM3:AQ3", "缺勤(天)"),
        ("AR3:AT3", "加班工时\n(H)"),
        ("AU3:AV3", "上月余加班工时\n(H)"),
        ("AW3:AX3", "剩余加班工时\n(H)"),
    ]
    for merge_range, title in stats_merges_row3:
        ws.merge_cells(merge_range)
        cell = ws[merge_range.split(":")[0]]
        cell.value = title
        align = unit_bottom_align if "\n(H)" in title else center_align
        _style_cell(cell, header_font, align, full_border)
        _apply_merge_range_borders(ws, merge_range)

    stats_merges_vertical = [
        ("AY3:AY4", "合计天数"),
        ("AZ3:AZ4", "夜班餐补天数"),
        ("BA3:BA4", "剩余年假\n(H)"),
        ("BB3:BB4", "确认签名"),
        ("BC3:BC4", "备注"),
        ("BD3:BD4", "入职日期"),
        ("BE3:BE4", "工龄"),
    ]
    for merge_range, title in stats_merges_vertical:
        ws.merge_cells(merge_range)
        top_left = merge_range.split(":")[0]
        cell = ws[top_left]
        cell.value = title
        _style_cell(cell, header_font, vertical_align, full_border)
        _apply_merge_range_borders(ws, merge_range)

    row4_headers = {
        35: "实际出勤",
        36: "调休\n(h)",
        37: "年休假\n(h)",
        38: "合计",
        39: "事假",
        40: "病假",
        41: "婚丧产",
        42: "工伤",
        43: "旷工",
        44: "平时",
        45: "休息日",
        46: "节假日",
        47: "平时",
        48: "周六日",
        49: "平时",
        50: "周六日",
    }
    for col, title in row4_headers.items():
        cell = ws.cell(4, col, title)
        _style_cell(cell, header_font, vertical_align, full_border)

    for col_letter, width in STATS_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def _merge_stats_cells_for_employee(
    ws, start_row, end_row, normal_font, center_align, col_values=None
):
    """右侧统计栏与序号/姓名一致：每位员工占两行，每列纵向合并。"""
    col_values = col_values or {}
    for col in range(STATS_COL_START, STATS_COL_END + 1):
        ws.merge_cells(
            start_row=start_row,
            start_column=col,
            end_row=end_row,
            end_column=col,
        )
        cell = ws.cell(start_row, col)
        if col in col_values:
            cell.value = col_values[col]
        _style_cell(cell, normal_font, center_align, _full_thin_border())
        _apply_vertical_merge_borders(ws, col, start_row, end_row)


def _employee_stats_values(check_count, weekday_ot=0.0, rest_ot=0.0):
    """员工统计区当前已实现的字段。"""
    return {
        COL_STATS_ACTUAL: check_count,
        COL_STATS_COMP_LEAVE: 0,
        COL_STATS_ANNUAL_LEAVE: 0,
        COL_STATS_ATTEND_SUM: check_count,
        COL_STATS_PERSONAL_LEAVE: 0,
        COL_STATS_SICK_LEAVE: 0,
        COL_STATS_FAMILY_LEAVE: 0,
        COL_STATS_WORK_INJURY: 0,
        COL_STATS_ABSENT: 0,
        COL_STATS_OT_WEEKDAY: _stats_hours_value(weekday_ot),
        COL_STATS_OT_REST: _stats_hours_value(rest_ot),
        COL_STATS_OT_HOLIDAY: 0,
    }


def generate_attendance_report():
    """生成考勤统计表"""
    
    # 读取原始打卡数据
    try:
        punch_df = pd.read_excel(PUNCH_FILE, engine='xlrd')
    except:
        try:
            punch_df = pd.read_excel(PUNCH_FILE, engine='openpyxl')
        except Exception as e:
            print(f"读取打卡文件失败: {e}")
            return
    
    punch_df['日期时间'] = pd.to_datetime(punch_df['日期时间'])
    punch_df['日期'] = punch_df['日期时间'].dt.date
    punch_df['打卡时间'] = punch_df['日期时间'].dt.time

    six_punch_names = refresh_workday_six_punch_employees_from_df(punch_df)
    print(f"\n六次卡工作日加班名单（来自四次基本卡部门，共 {len(six_punch_names)} 人）:")
    print(sorted(six_punch_names))
    
    # 读取异常记录（如果文件不存在或读取失败，创建空DataFrame）
    anomaly_df = pd.DataFrame()
    if os.path.exists(ANOMALY_FILE):
        try:
            anomaly_df = pd.read_excel(ANOMALY_FILE)
            anomaly_df['日期'] = pd.to_datetime(anomaly_df['日期']).dt.date
        except Exception as e:
            print(f"读取异常文件失败（将继续生成不含异常标记的表格）: {e}")
    else:
        print(f"异常文件不存在: {ANOMALY_FILE}，将生成不含异常标记的表格")
    
    # 获取日期范围和员工列表
    start_date = punch_df['日期'].min()
    end_date = punch_df['日期'].max()
    month = start_date.month
    year = start_date.year
    days_in_month = pd.Period(f'{year}-{month}').days_in_month
    
    employees = punch_df[['姓名', '编号']].drop_duplicates()
    agency_keys = get_agency_employee_keys(punch_df)

    # 计算工作日和休息日
    workdays = sum(1 for day in range(1, days_in_month + 1) 
                   if is_workday(datetime(year, month, day).date()))
    rest_days = days_in_month - workdays
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}月"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    # 设置日期列宽度（D到AH列，共31列）
    date_cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH']
    for col in date_cols:
        ws.column_dimensions[col].width = 4
    
    # 样式定义
    gray_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅米色阴影
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(name='黑体', size=18, bold=True)
    normal_font = Font(name='宋体', size=10)
    header_font = Font(name='宋体', size=11)
    
    # 第1行：大标题
    ws.merge_cells('A1:BE1')
    ws['A1'] = f"{year}年{month}月份考勤确认表"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws.row_dimensions[1].height = 30
    
    # 第2行：单位信息（不合并，手动分三段）
    ws['A2'] = "单位:珠海亚拓生物科技有限公司"
    ws['A2'].font = header_font
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells('K2:R2')
    ws['K2'] = f"全勤天数{workdays}天，休{rest_days}天"
    ws['K2'].font = header_font
    ws['K2'].alignment = center_align
    
    ws.merge_cells('X2:BE2')
    ws['X2'] = "注：阴影部份为周六日出勤情况"
    ws['X2'].font = header_font
    ws['X2'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws.row_dimensions[2].height = 20
    
    # 第3-4行：表头
    vertical_align = _vertical_header_align()
    ws.merge_cells('A3:A4')
    ws['A3'] = "序号"
    ws['A3'].alignment = vertical_align
    ws['A3'].font = header_font
    _apply_merge_range_borders(ws, 'A3:A4')

    ws.merge_cells('B3:B4')
    ws['B3'] = "        日期\n\n 姓名"
    ws['B3'].alignment = center_align
    ws['B3'].font = header_font
    _apply_merge_range_borders(ws, 'B3:B4')

    ws.merge_cells('C3:C4')
    ws['C3'] = ""
    _apply_merge_range_borders(ws, 'C3:C4')
    
    ws.row_dimensions[3].height = 47.25
    ws.row_dimensions[4].height = 69.75
    
    # 冻结前4行表头 + 前3列(A-C：序号/姓名/行标签)，滚动从 D5 开始
    ws.freeze_panes = "D5"
    
    # 写入日期列（1-31），从D列开始
    for day in range(1, days_in_month + 1):
        col_idx = 3 + day  # D列是第4列
        ws.merge_cells(start_row=3, start_column=col_idx, end_row=4, end_column=col_idx)
        cell = ws.cell(3, col_idx, day)
        cell.alignment = center_align
        cell.font = normal_font
        _apply_vertical_merge_borders(ws, col_idx, 3, 4)

        # 节假日/周末加阴影
        date = datetime(year, month, day).date()
        if is_holiday_or_weekend(date):
            cell.fill = gray_fill

    _setup_stats_headers(ws, workdays, header_font, center_align)
    
    # 填充员工数据，从第5行开始
    current_row = 5
    seq_num = 1
    
    for _, emp_row in employees.iterrows():
        name = emp_row['姓名']
        emp_id = emp_row['编号']
        
        emp_records = punch_df[punch_df['姓名'] == name]
        emp_punch_dates = set(emp_records['日期'])
        # 按日期汇总打卡时间（已排序）
        emp_punches_by_date = {}
        for date, group in emp_records.groupby('日期'):
            emp_punches_by_date[date] = sorted(group['打卡时间'].tolist())

        month_start = datetime(year, month, 1).date()
        month_end = datetime(year, month, days_in_month).date()

        # 获取该员工的异常记录（保留完整信息用于批注）
        emp_anomalies = {}
        emp_anomaly_details = {}  # 保存详细信息
        if not anomaly_df.empty:
            emp_anomaly_records = anomaly_df[anomaly_df['姓名'] == name]
            for _, anomaly in emp_anomaly_records.iterrows():
                date = anomaly['日期']
                if date not in emp_anomalies:
                    emp_anomalies[date] = []
                    emp_anomaly_details[date] = []
                emp_anomalies[date].append(anomaly['考勤异常情况'])
                emp_anomaly_details[date].append({
                    '时间': anomaly.get('打卡时间', ''),
                    '异常': anomaly['考勤异常情况']
                })

        missing_night_start = {}
        if name in workday_overtime.TARGET_WORKDAY_SIX_PUNCH_EMPLOYEES:
            missing_night_start = collect_missing_night_start_dates(
                emp_punches_by_date, month_start, month_end
            )
            for missing_date, info in missing_night_start.items():
                if missing_date not in emp_anomalies:
                    emp_anomalies[missing_date] = []
                    emp_anomaly_details[missing_date] = []
                if "缺卡" not in emp_anomalies[missing_date]:
                    emp_anomalies[missing_date].append("缺卡")
                    emp_anomaly_details[missing_date].append({
                        "时间": str(info["ref_time"]),
                        "异常": "缺卡",
                    })
        
        emp_start_row = current_row

        # 设置行高
        ws.row_dimensions[current_row].height = 20
        ws.row_dimensions[current_row + 1].height = 20
        
        # 序号（跨两行）
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=1)
        cell_seq = ws.cell(current_row, 1, seq_num)
        cell_seq.alignment = center_align
        cell_seq.font = normal_font
        _apply_vertical_merge_borders(ws, 1, current_row, current_row + 1)

        # 姓名（跨两行）
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+1, end_column=2)
        cell_name = ws.cell(current_row, 2, name)
        cell_name.alignment = center_align
        cell_name.font = normal_font
        _apply_vertical_merge_borders(ws, 2, current_row, current_row + 1)
        
        # 第1行：正常出勤
        cell_label = ws.cell(current_row, 3, "正常出勤")
        cell_label.alignment = center_align
        cell_label.font = normal_font
        cell_label.border = thin_border
        
        # 填充每日考勤，从D列开始
        check_count = 0
        for day in range(1, days_in_month + 1):
            col_idx = 3 + day
            date = datetime(year, month, day).date()
            
            cell = ws.cell(current_row, col_idx)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center_align
            
            # 节假日/周末加阴影
            if is_holiday_or_weekend(date):
                cell.fill = gray_fill
            
            # 判断是否有异常
            if date in emp_anomalies:
                # 如果异常类型包含"缺勤"，显示"缺"
                if any('缺勤' in a for a in emp_anomalies[date]):
                    cell.value = "缺"
                else:
                    cell.value = "异"
                
                # 添加批注显示详细异常信息
                comment_lines = []
                for detail in emp_anomaly_details[date]:
                    time_str = detail['时间']
                    anomaly_type = detail['异常']
                    comment_lines.append(f"{date.strftime('%Y/%m/%d')}\n{time_str}\n{anomaly_type}")
                cell.comment = Comment('\n\n'.join(comment_lines), "系统")
                
            elif date in emp_punch_dates and not is_holiday_or_weekend(date):
                cell.value = "√"
                check_count += 1
            elif is_workday(date):
                # 工作日无打卡=缺勤
                cell.value = "缺"
                # 添加批注
                cell.comment = Comment(f"{date.strftime('%Y/%m/%d')}\n\n缺勤", "系统")

        # 第2行：加班工时
        current_row += 1
        cell_label2 = ws.cell(current_row, 3, "加班工时")
        cell_label2.alignment = center_align
        cell_label2.font = normal_font
        cell_label2.border = thin_border
        
        # 加班工时行
        weekday_ot_total = 0.0
        rest_ot_total = 0.0
        is_agency = (name, emp_id) in agency_keys
        for day in range(1, days_in_month + 1):
            col_idx = 3 + day
            date = datetime(year, month, day).date()
            cell = ws.cell(current_row, col_idx)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center_align
            if is_holiday_or_weekend(date):
                cell.fill = gray_fill

            if is_agency:
                # 中介：早班/夜班工时统一计入「平时」
                hours = calc_agency_hours_for_shift_end_date(
                    emp_punches_by_date, date
                )
                if hours is not None and hours > 0:
                    cell.value = _format_overtime_hours(hours)
                    weekday_ot_total += hours
                continue

            if is_holiday_or_weekend(date):
                day_times = emp_punches_by_date.get(date, [])
                if name in workday_overtime.TARGET_WORKDAY_SIX_PUNCH_EMPLOYEES:
                    result = calc_four_punch_department_rest_overtime(
                        date, day_times, emp_punches_by_date
                    )
                    if (
                        result["status"] == "正常"
                        and result["hours"] is not None
                        and result["hours"] > 0
                    ):
                        cell.value = _format_overtime_hours(result["hours"])
                        rest_ot_total += result["hours"]
                    elif result["status"] == "异常":
                        cell.value = "异"
                    elif date in missing_night_start:
                        cell.value = "异"
                        info = missing_night_start[date]
                        cell.comment = Comment(
                            f"{date.strftime('%Y/%m/%d')}\n{info['ref_time']}\n缺卡",
                            "系统",
                        )
                elif len(day_times) == 2:
                    hours = calc_weekend_overtime_two_punches(
                        day_times[0], day_times[1]
                    )
                    if hours is not None and hours > 0:
                        cell.value = _format_overtime_hours(hours)
                        rest_ot_total += hours
            else:
                # 工作日加班：仅统计指定名单员工
                if (
                    name in TARGET_WORKDAY_OVERTIME_EMPLOYEES
                    or name in workday_overtime.TARGET_WORKDAY_SIX_PUNCH_EMPLOYEES
                ):
                    day_times = emp_punches_by_date.get(date, [])
                    result = calc_workday_overtime(
                        name, date, emp_id, day_times, emp_punches_by_date
                    )
                    if result["status"] == "正常" and result["hours"] is not None and result["hours"] > 0:
                        cell.value = _format_overtime_hours(result["hours"])
                        weekday_ot_total += result["hours"]
                    elif result["status"] == "异常":
                        cell.value = "异"
                    elif date in missing_night_start:
                        cell.value = "异"
                        info = missing_night_start[date]
                        cell.comment = Comment(
                            f"{date.strftime('%Y/%m/%d')}\n{info['ref_time']}\n缺卡",
                            "系统",
                        )

        stats_values = _employee_stats_values(check_count, weekday_ot_total, rest_ot_total)
        _merge_stats_cells_for_employee(
            ws, emp_start_row, current_row, normal_font, center_align, stats_values
        )

        current_row += 1
        seq_num += 1

    # 与模板一致：第4行表头启用自动筛选（升序/降序下拉箭头）
    last_data_row = current_row - 1
    if last_data_row >= 5:
        ws.auto_filter.ref = f"A4:BE{last_data_row}"

    # 保存文件
    output_file = os.path.join(OUTPUT_DIR, f"{month}月考勤统计.xlsx")
    wb.save(output_file)
    print(f"考勤统计表已生成: {output_file}")
    print(f"共处理 {len(employees)} 名员工")


if __name__ == "__main__":
    generate_attendance_report()
